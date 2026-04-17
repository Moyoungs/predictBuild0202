"""
parser.py — Excel 집계표 시트 파싱

책임:
  1. 파일의 시트 중 '집계표' 시트 식별
  2. 헤더 행과 데이터 영역 탐지 (다양한 레이아웃 지원)
  3. 집계(소계/합계) 행 제외
  4. 구조화된 데이터(레코드 리스트) 반환

지원 레이아웃:
  A) 공종(0) | 규격1(1) | 규격2(2) | 단위(3) | 구조물1(4) | 구조물2(5) | 계 | 비고
     (01, 02, 03, 04번 파일)
  B) 번호(0) | 공종(1) | 규격(2) | 단위(3) | 수량(4) | 비고(5)
     (05번 파일)
"""
from pathlib import Path
from typing import Optional
import warnings

from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=UserWarning)

# 구조물 컬럼에서 제외할 키워드 (수량 컬럼이 아님)
NON_DATA_COLUMNS = {"계", "합계", "소계", "총계", "비고", "Total", "TOTAL", "total",
                    "Sub Total", "합  계", "합    계", "소  계", "소    계",
                    "도면에서 확인 후 입력", "", "None"}


def classify_sheet(sheet_name: str, terminology: dict) -> str:
    """시트명을 보고 용도 분류: aggregate | calculation | cover | ignore"""
    sc = terminology["sheet_classification"]
    name = sheet_name.strip()

    # ignore 먼저 (Sheet1 등)
    for kw in sc["ignore"]["keywords"]:
        if name == kw:
            return "ignore"

    # cover/간지를 aggregate보다 먼저 체크 (표지가 "집계표"보다 우선)
    for kw in sc["cover_sheet"]["keywords"]:
        if kw in name:
            return "cover"

    # 그 다음 집계표
    for kw in sc["aggregate_sheet"]["keywords"]:
        if kw in name:
            return "aggregate"

    # 산출근거
    for kw in sc["calculation_sheet"]["keywords"]:
        if kw in name:
            return "calculation"

    return "ignore"

def find_aggregate_sheets(wb, terminology: dict) -> list[str]:
    """
    워크북에서 집계표 시트만 골라내기.
    규칙:
      - '집계' 키워드가 있는 시트를 후보로 수집
      - 'prefer_patterns'에 매칭되는 시트가 하나라도 있으면
        'partial_keywords'가 포함된 시트는 제외 (중복 집계 방지)
    """
    sc = terminology["sheet_classification"]["aggregate_sheet"]
    prefer = sc.get("prefer_patterns", [])
    partial = sc.get("partial_keywords", [])

    candidates = [s for s in wb.sheetnames if classify_sheet(s, terminology) == "aggregate"]

    # prefer 매칭 여부 판정
    has_full = any(any(p in s for p in prefer) for s in candidates)

    if has_full:
        # partial만 있는 시트 제거
        result = []
        for s in candidates:
            is_partial = any(p in s for p in partial)
            is_full = any(p in s for p in prefer)
            # full에 매칭되면 포함, partial인데 full 매칭 아니면 제외
            if is_full or not is_partial:
                result.append(s)
        return result
    return candidates


def _norm(s: str) -> str:
    """공백/개행 전부 제거"""
    return "".join(str(s).split()) if s is not None else ""


def find_header_row(ws) -> Optional[int]:
    """
    헤더 행 탐지:
      - '공종' 또는 '단위' 키워드 포함 행
      - 첫 10행 이내
    """
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        for cell in ws[row_idx]:
            if cell.value is None: continue
            v = _norm(cell.value)
            if v == "공종" or v == "단위":
                return row_idx
    return None


def find_column_indexes(ws, header_row: int) -> dict:
    """
    헤더 행에서 주요 컬럼 위치 찾기:
      - work_type_col: '공종'
      - unit_col: '단위'
      - data_cols: 단위 다음 컬럼부터, NON_DATA_COLUMNS 제외 + (컬럼명, 인덱스) 튜플
    병합 셀로 인해 헤더가 header_row+1에 있을 수도 있음 → 해당 행도 확인
    """
    header_cells = list(ws[header_row])
    # 병합된 헤더 대응: header_row+1에 세부 헤더가 있으면 병합
    next_row_cells = list(ws[header_row + 1]) if ws.max_row > header_row else []

    def header_value_at(col_idx):
        """컬럼 인덱스의 헤더 값 (두 줄 헤더 대응)"""
        v1 = header_cells[col_idx].value if col_idx < len(header_cells) else None
        v2 = next_row_cells[col_idx].value if col_idx < len(next_row_cells) else None
        return v1 or v2

    work_type_col = None
    unit_col = None
    max_col = max(len(header_cells), len(next_row_cells))

    for idx in range(max_col):
        v = header_value_at(idx)
        if v is None: continue
        n = _norm(v)
        if work_type_col is None and n == "공종":
            work_type_col = idx
        if unit_col is None and n == "단위":
            unit_col = idx

    if unit_col is None:
        return {}

    # 단위 이후 컬럼 = 구조물별 수량 컬럼 후보
    data_cols = []
    for idx in range(unit_col + 1, max_col):
        v = header_value_at(idx)
        if v is None: continue
        v_str = str(v).strip()
        if not v_str: continue
        if v_str in NON_DATA_COLUMNS: continue
        data_cols.append((idx, v_str))

    return {
        "work_type_col": work_type_col if work_type_col is not None else 0,
        "unit_col": unit_col,
        "data_cols": data_cols,
        "header_row": header_row,
    }


def is_aggregate_row(row_values: tuple, terminology: dict) -> bool:
    keywords = terminology["aggregate_row_keywords"]
    normalized = [_norm(k) for k in keywords]
    for v in row_values:
        if v is None: continue
        if _norm(v) in normalized:
            return True
    return False


def is_empty_value(v, terminology: dict) -> bool:
    if v is None: return True
    s = str(v).strip()
    return s in terminology["empty_values"] or s == ""


def parse_numeric(v) -> Optional[float]:
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip()
    if s in ("-", "", "N/A", "해당없음"): return 0.0
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def parse_aggregate_sheet(wb, sheet_name: str, terminology: dict, source_file: str) -> list[dict]:
    ws = wb[sheet_name]
    header_row = find_header_row(ws)
    if header_row is None:
        return []

    cols = find_column_indexes(ws, header_row)
    if not cols:
        return []

    work_type_col = cols["work_type_col"]
    unit_col = cols["unit_col"]
    data_cols = cols["data_cols"]

    # 헤더가 2줄이면 데이터 시작이 header_row+2
    # 간단 판정: header_row+1 행에 '계', '본체' 등 서브헤더가 있는지 확인
    data_start = header_row + 1
    next_row_vals = [c.value for c in ws[header_row + 1]] if ws.max_row > header_row else []
    if any(_norm(v) in ("계", "본체", "시점측", "종점측", "시점", "종점") for v in next_row_vals if v):
        data_start = header_row + 2

    records = []
    current_work_type = None
    last_spec1 = ""  # 병합 셀 대응: 규격1이 비면 이전 값 상속

    for row_idx, row in enumerate(
        ws.iter_rows(values_only=True, min_row=data_start), start=data_start
    ):
        if is_aggregate_row(row, terminology):
            continue

        # 공종 컬럼 (병합으로 빈칸이면 이전 값 유지)
        col_wt = row[work_type_col] if len(row) > work_type_col else None
        wt_changed = False
        if col_wt is not None and not is_empty_value(col_wt, terminology):
            v = str(col_wt).strip()
            if _norm(v) not in [_norm(k) for k in terminology["aggregate_row_keywords"]]:
                if current_work_type != v:
                    wt_changed = True
                    last_spec1 = ""  # 공종 바뀌면 규격 상속 리셋
                current_work_type = v

        # 규격 (work_type_col + 1, work_type_col + 2)
        spec1_idx = work_type_col + 1
        spec2_idx = work_type_col + 2
        spec1 = row[spec1_idx] if len(row) > spec1_idx and spec1_idx != unit_col else None
        spec2 = row[spec2_idx] if len(row) > spec2_idx and spec2_idx != unit_col else None

        # 규격1 상속: 같은 공종 내에서 규격1이 비어있으면 이전 값 사용
        spec1_str = str(spec1).strip() if spec1 else ""
        if spec1_str:
            last_spec1 = spec1_str
        else:
            spec1_str = last_spec1  # 상속

        # 단위
        unit_raw = row[unit_col] if len(row) > unit_col else None

        # 수량 수집
        quantities = {}
        for col_idx, col_name in data_cols:
            if col_idx >= len(row): continue
            val = row[col_idx]
            num = parse_numeric(val)
            if num is not None and num != 0:
                quantities[col_name] = num

        if not current_work_type: continue
        if not quantities: continue

        records.append({
            "source_file": source_file,
            "sheet_name": sheet_name,
            "row_index": row_idx,
            "work_type_raw": current_work_type,
            "spec1_raw": spec1_str,
            "spec2_raw": str(spec2).strip() if spec2 else "",
            "unit_raw": str(unit_raw).strip() if unit_raw else "",
            "quantities": quantities,
        })

    return records


def parse_workbook(path: Path, terminology: dict) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    agg_sheets = find_aggregate_sheets(wb, terminology)
    all_records = []
    for sheet_name in agg_sheets:
        records = parse_aggregate_sheet(wb, sheet_name, terminology, source_file=path.name)
        all_records.extend(records)
    return all_records


if __name__ == "__main__":
    import yaml
    from pathlib import Path

    # 프로젝트 루트 경로 (이 파일 기준으로 3단계 위)
    project_root = Path(__file__).resolve().parent.parent.parent.parent
    terminology_path = project_root / "data" / "terminology.yaml"
    fixtures_path = project_root / "tests" / "fixtures" / "galjeon7"

    with open(terminology_path, encoding="utf-8") as f:
        term = yaml.safe_load(f)

    # fixtures 폴더에서 Excel 파일 자동 탐색
    files = sorted(fixtures_path.glob("*.xlsx")) + sorted(fixtures_path.glob("*.xls"))
    # 00번(총괄집계표)은 출력 대상이므로 제외
    files = [f for f in files if not f.name.startswith("00_")]

    if not files:
        print(f"❌ 테스트 파일이 없습니다: {fixtures_path}")
        print("   tests/fixtures/galjeon7/ 에 Excel 파일을 넣어주세요.")
        exit(1)

    for f in files:
        wb = load_workbook(f, data_only=True) if f.suffix.lower() != '.xls' else None
        if wb is None:
            xlsx_f = f.with_suffix('.xlsx')
            if xlsx_f.exists():
                wb = load_workbook(xlsx_f, data_only=True)
        if wb:
            from parser import find_aggregate_sheets
            agg = find_aggregate_sheets(wb, term)
            print(f"\n{f.name}:")
            print(f"  전체 시트: {wb.sheetnames}")
            print(f"  파싱 대상: {agg}")

    all_records = []
    for f in files:
        records = parse_workbook(f, term)
        print(f"{f.name}: {len(records)}개 레코드")
        all_records.extend(records)

    print(f"\n총 레코드: {len(all_records)}")
    print("\n[각 파일 첫 레코드]")
    seen = set()
    for r in all_records:
        if r['source_file'] not in seen:
            seen.add(r['source_file'])
            print(f"\n  {r['source_file']} / {r['sheet_name']}")
            print(f"    공종: {r['work_type_raw']}")
            print(f"    규격: {r['spec1_raw']} / {r['spec2_raw']}")
            print(f"    단위: {r['unit_raw']}")
            print(f"    수량: {r['quantities']}")