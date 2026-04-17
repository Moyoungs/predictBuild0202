"""
reporter.py — 집계 결과를 Excel 파일로 출력

MVP v1: 간단한 3개 시트 구조
  - 구조물공총괄집계표
  - 토공총괄집계표
  - 철근가공조립 (Type별)
  + 검증리포트 시트

추후: 00번 템플릿 복제 방식으로 고도화 예정
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def _header_style():
    return {
        "font": Font(bold=True, size=11),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "fill": PatternFill("solid", fgColor="D9E1F2"),
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        ),
    }


def _cell_style():
    return {
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        ),
    }


def _apply_style(cell, style_dict):
    for k, v in style_dict.items():
        setattr(cell, k, v)


def _get_all_structures(agg_rows: list[dict]) -> list[str]:
    """집계 결과에서 등장한 모든 구조물명 추출"""
    structures = set()
    for row in agg_rows:
        structures.update(row["structures"].keys())
    # 관례적인 정렬: 본교 먼저, 그 다음 옹벽, 가축도, 기타
    priority = ["본교", "옹벽", "가축도"]
    ordered = [s for s in priority if s in structures]
    ordered += sorted(structures - set(priority))
    return ordered


def write_aggregate_sheet(ws, title: str, agg_rows: list[dict]):
    """집계 시트 하나 생성"""
    structures = _get_all_structures(agg_rows)

    # 제목
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.row_dimensions[1].height = 24

    # 헤더 (3행)
    header_row = 3
    headers = ["공종", "규격1", "규격2", "단위"] + structures + ["총계", "비고"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        _apply_style(cell, _header_style())

    # 데이터
    for r_idx, row in enumerate(agg_rows, start=header_row + 1):
        ws.cell(row=r_idx, column=1, value=row["work_type"])
        ws.cell(row=r_idx, column=2, value=row["spec1"])
        ws.cell(row=r_idx, column=3, value=row["spec2"])
        ws.cell(row=r_idx, column=4, value=row["unit"])
        for s_idx, struct in enumerate(structures, start=5):
            val = row["structures"].get(struct, 0)
            cell = ws.cell(row=r_idx, column=s_idx, value=val if val > 0 else None)
            if val > 0:
                cell.number_format = "#,##0.000"
        # 총계
        total_cell = ws.cell(row=r_idx, column=5 + len(structures), value=row["total"])
        total_cell.number_format = "#,##0.000"
        total_cell.font = Font(bold=True)

        # 스타일 적용
        for c in range(1, 5 + len(structures) + 2):
            _apply_style(ws.cell(row=r_idx, column=c), _cell_style())

    # 컬럼 너비
    widths = [15, 16, 20, 8] + [12] * len(structures) + [14, 10]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def write_validation_sheet(ws, validation_results: list[dict]):
    """검증 결과 시트 (원본 vs 프로토타입 비교)"""
    ws.cell(row=1, column=1, value="검증 리포트 (원본 00번 총괄집계표 대비)").font = Font(bold=True, size=14)

    headers = ["공종", "규격1", "규격2", "원본 값", "생성 값", "차이", "판정"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        _apply_style(cell, _header_style())

    for r_idx, v in enumerate(validation_results, start=4):
        ws.cell(row=r_idx, column=1, value=v["work_type"])
        ws.cell(row=r_idx, column=2, value=v["spec1"])
        ws.cell(row=r_idx, column=3, value=v["spec2"])
        ws.cell(row=r_idx, column=4, value=v["expected"]).number_format = "#,##0.000"
        ws.cell(row=r_idx, column=5, value=v["actual"]).number_format = "#,##0.000"
        ws.cell(row=r_idx, column=6, value=v["diff"]).number_format = "#,##0.000"
        ws.cell(row=r_idx, column=7, value="일치" if v["ok"] else "불일치")

        for c in range(1, 8):
            _apply_style(ws.cell(row=r_idx, column=c), _cell_style())

    widths = [15, 15, 20, 12, 12, 10, 10]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def write_unmatched_sheet(ws, records: list[dict]):
    """미매칭 항목(AI 판단 필요) 시트"""
    ws.cell(row=1, column=1, value="미매칭 항목 (용어 사전 보강 필요)").font = Font(bold=True, size=14)

    headers = ["파일", "시트", "공종(원본)", "규격1", "단위", "사유"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        _apply_style(cell, _header_style())

    r_idx = 4
    for r in records:
        n = r["normalized"]
        if n["unmatched"]:
            for reason in n["unmatched"]:
                ws.cell(row=r_idx, column=1, value=r["source_file"])
                ws.cell(row=r_idx, column=2, value=r["sheet_name"])
                ws.cell(row=r_idx, column=3, value=r["work_type_raw"])
                ws.cell(row=r_idx, column=4, value=r["spec1_raw"])
                ws.cell(row=r_idx, column=5, value=r["unit_raw"])
                ws.cell(row=r_idx, column=6, value=reason)
                for c in range(1, 7):
                    _apply_style(ws.cell(row=r_idx, column=c), _cell_style())
                r_idx += 1

    widths = [35, 25, 15, 20, 8, 40]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def generate_report(
    output_path: str,
    structure_agg: list[dict],
    earthwork_agg: list[dict],
    temp_road_agg: list[dict],
    ground_imp_agg: list[dict],
    temp_struct_agg: list[dict],
    rebar_agg: dict,
    all_records: list[dict],
    validation_results: list[dict] = None,
):
    wb = Workbook()
    # 기본 시트 제거
    wb.remove(wb.active)

    # 구조물공 총괄
    ws = wb.create_sheet("구조물공총괄집계표")
    write_aggregate_sheet(ws, "구조물공 총괄집계표", structure_agg)

    # 토공 총괄
    ws = wb.create_sheet("토공총괄집계표")
    write_aggregate_sheet(ws, "토공 총괄집계표", earthwork_agg)

    # 가시설공 (보너스)
    if temp_struct_agg:
        ws = wb.create_sheet("가시설공총괄")
        write_aggregate_sheet(ws, "가시설공 총괄집계표", temp_struct_agg)

    # 가축도 (보너스)
    if temp_road_agg:
        ws = wb.create_sheet("가축도총괄")
        write_aggregate_sheet(ws, "가축도 총괄집계표", temp_road_agg)

    # 지반개량 (보너스)
    if ground_imp_agg:
        ws = wb.create_sheet("지반개량총괄")
        write_aggregate_sheet(ws, "지반개량공 총괄집계표", ground_imp_agg)

    # 철근가공조립 Type별
    ws = wb.create_sheet("철근가공조립")
    ws.cell(row=1, column=1, value="철근가공조립 Type별 집계").font = Font(bold=True, size=14)
    headers = ["Type", "본교", "옹벽", "총계(ton)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        _apply_style(cell, _header_style())
    seen = set()
    r_idx = 4
    for r in rebar_agg["by_type"]:
        if r["total"] <= 0: continue
        if r["type"] in seen: continue
        seen.add(r["type"])
        ws.cell(row=r_idx, column=1, value=r["type"])
        ws.cell(row=r_idx, column=2, value=r["structures"].get("본교", 0)).number_format = "#,##0.000"
        ws.cell(row=r_idx, column=3, value=r["structures"].get("옹벽", 0)).number_format = "#,##0.000"
        ws.cell(row=r_idx, column=4, value=r["total"]).number_format = "#,##0.000"
        for c in range(1, 5):
            _apply_style(ws.cell(row=r_idx, column=c), _cell_style())
        r_idx += 1
    for idx, w in enumerate([15, 12, 12, 14], 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # 검증 리포트
    if validation_results:
        ws = wb.create_sheet("검증리포트")
        write_validation_sheet(ws, validation_results)

    # 미매칭
    ws = wb.create_sheet("미매칭항목")
    write_unmatched_sheet(ws, all_records)

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    import yaml
    from pathlib import Path
    from parser import parse_workbook
    from normalizer import normalize_all
    from aggregator import aggregate_by_category, aggregate_rebar

    project_root = Path(__file__).resolve().parent.parent.parent.parent
    terminology_path = project_root / "data" / "terminology.yaml"
    fixtures_path = project_root / "tests" / "fixtures" / "galjeon7"

    with open(terminology_path, encoding="utf-8") as f:
        term = yaml.safe_load(f)

    files = sorted(fixtures_path.glob("*.xlsx")) + sorted(fixtures_path.glob("*.xls"))
    files = [f for f in files if not f.name.startswith(("00_", "00 "))]

    if not files:
        print(f"❌ 테스트 파일이 없습니다: {fixtures_path}")
        exit(1)

    all_records = []
    for f in files:
        all_records.extend(parse_workbook(f, term))
    normalize_all(all_records, term)

    structure_agg = aggregate_by_category(all_records, "structure", term)
    earthwork_agg = aggregate_by_category(all_records, "earthwork", term)
    temp_road_agg = aggregate_by_category(all_records, "temp_road", term)
    ground_imp_agg = aggregate_by_category(all_records, "ground_improvement", term)
    temp_struct_agg = aggregate_by_category(all_records, "temp_structure", term)
    rebar_agg = aggregate_rebar(all_records)

    # 검증 데이터
    actual = {(r["work_type"], r["spec1"], r["spec2"]): r["total"] for r in structure_agg}
    concrete_35_total = sum(
        v for (wt, s1, s2), v in actual.items()
        if wt == "콘크리트" and s1 == "25-35-15"
    )

    expected = [
        ("콘크리트", "25-35-15", "본체", 963.067),
        ("콘크리트", "25-27-15", "기초", 464.94),
        ("콘크리트", "25-18-15", "버림콘크리트", 81.368),
        ("콘크리트타설", "펌프카타설", "0~15m", 1428.007),
        ("거푸집", "합판4회", "", 38.465),
        ("거푸집", "합판6회", "수직면(0~7m)", 28.209),
        ("동바리", "시스템동바리", "0.0~10.0m", 2245.057),
        ("ASP방수", "", "", 672.474),
        ("TBM설치", "", "", 1.0),
        ("교명판", "", "", 1.0),
    ]

    earth_actual = {(r["work_type"], r["spec1"], r["spec2"]): r["total"] for r in earthwork_agg}
    earth_expected = [
        ("되메우기", "", "", 3818.185),
        ("뒷채움", "", "", 1199.905),
        ("물푸기", "", "", 433.572),
        ("유용토", "", "", 1375.545),
        ("터파기", "토사", "(0~4m)", 1748.131),
    ]

    validation = []
    for wt, s1, s2, exp in expected:
        if (wt, s1, s2) == ("콘크리트", "25-35-15", "본체"):
            act = concrete_35_total
        else:
            act = actual.get((wt, s1, s2), 0.0)
        validation.append({
            "work_type": wt, "spec1": s1, "spec2": s2,
            "expected": exp, "actual": act,
            "diff": round(abs(act - exp), 4),
            "ok": abs(act - exp) < 0.01,
        })
    for wt, s1, s2, exp in earth_expected:
        act = earth_actual.get((wt, s1, s2), 0.0)
        validation.append({
            "work_type": wt, "spec1": s1, "spec2": s2,
            "expected": exp, "actual": act,
            "diff": round(abs(act - exp), 4),
            "ok": abs(act - exp) < 0.01,
        })

    # 출력 폴더
    output_dir = project_root / "outputs"
    output_dir.mkdir(exist_ok=True)
    output_path = str(output_dir / "갈전7교_자동생성_총괄집계표.xlsx")

    generate_report(
        output_path,
        structure_agg, earthwork_agg,
        temp_road_agg, ground_imp_agg, temp_struct_agg,
        rebar_agg, all_records, validation
    )

    hit = sum(1 for v in validation if v["ok"])
    print(f"✅ 생성 완료: {output_path}")
    print(f"   검증 정확도: {hit}/{len(validation)} = {hit/len(validation)*100:.1f}%")
    print("\n[검증 상세]")
    for v in validation:
        mark = "✅" if v["ok"] else "❌"
        print(f"  {mark} {v['work_type']:<12} {v['spec1']:<12} {v['spec2']:<15} "
              f"원본={v['expected']:>12,.3f}  생성={v['actual']:>12,.3f}  차이={v['diff']}")
    # 중복 원인 추적
    print("\n[중복 추적: ASP방수]")
    for r in all_records:
        n = r["normalized"]
        if n["work_type_std"] == "ASP방수":
            print(f"  파일: {r['source_file']}")
            print(f"  시트: {r['sheet_name']}")
            print(f"  수량: {r['quantities']}")
            print(f"  정규화: {n['quantities_std']}")
            print()
    print("[중복 추적: 물푸기]")
    for r in all_records:
        n = r["normalized"]
        if n["work_type_std"] == "물푸기":
            print(f"  파일: {r['source_file']}")
            print(f"  시트: {r['sheet_name']}")
            print(f"  수량: {r['quantities']}")
            print(f"  정규화: {n['quantities_std']}")
            print()